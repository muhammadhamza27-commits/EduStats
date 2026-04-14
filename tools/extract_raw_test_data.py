from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from statistics import mean
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - local dependency guard
    raise SystemExit("openpyxl is required. Install it with: pip install openpyxl") from exc


DEFAULT_INPUT_DIR = Path(r"c:\Users\Lenovo\Desktop\Code\RawTestData")
DEFAULT_OUTPUT_DIR = Path(r"c:\Users\Lenovo\Desktop\Code\Projects\generated-test-data")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def slugify(value: str) -> str:
    value = clean_text(value).lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-{2,}", "-", value).strip("-")
    return value or "unknown"


def to_number(value: Any) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return value
    text = clean_text(value)
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if number.is_integer():
        return int(number)
    return number


def format_csv_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.6f}".rstrip("0").rstrip(".")
    return str(value)


def ratio_to_percentage(value: Any) -> float | None:
    number = to_number(value)
    if number is None:
        return None
    if number <= 1:
        return round(float(number) * 100, 4)
    return round(float(number), 4)


def extract_first_number(value: Any) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    match = re.search(r"\d+(?:\.\d+)?", str(value))
    if not match:
        return None
    number = float(match.group(0))
    return int(number) if number.is_integer() else number


@dataclass
class SubjectRecord:
    name: str
    max_mark: float | int | None
    min_mark: float | int | None
    first_term: float | int | None
    mid_term: float | int | None
    final_term: float | int | None
    weighted_score: float | int | None
    overall_score: float | int | None
    total: float | int | None


@dataclass
class SheetRecord:
    workbook_file: str
    workbook_bucket: str
    class_label: str
    sheet_name: str
    academic_year: str
    student_id: str
    student_name: str
    father_name: str
    source_sheet_order: int
    subjects: list[SubjectRecord]
    total_max_marks: float | int | None
    total_min_marks: float | int | None
    class_total: float | int | None
    class_percentage: float | int | None
    attendance_count: float | int | None
    attendance_out_of: float | int | None
    attendance_percentage: float | int | None
    result: str
    grade: str
    position: str
    notes: str


def parse_subject_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[SubjectRecord]:
    subjects: list[SubjectRecord] = []
    for row in range(13, min(ws.max_row, 30) + 1):
        subject_name = clean_text(ws.cell(row, 2).value)
        if not subject_name:
            continue
        if subject_name.upper() in {"TOTAL", "PERCENTAGE"}:
            break

        subjects.append(
            SubjectRecord(
                name=subject_name,
                max_mark=to_number(ws.cell(row, 3).value),
                min_mark=to_number(ws.cell(row, 4).value),
                first_term=to_number(ws.cell(row, 5).value),
                mid_term=to_number(ws.cell(row, 6).value),
                final_term=to_number(ws.cell(row, 7).value),
                weighted_score=to_number(ws.cell(row, 8).value),
                overall_score=to_number(ws.cell(row, 9).value),
                total=to_number(ws.cell(row, 10).value),
            )
        )
    return subjects


def parse_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    workbook_file: str,
    workbook_bucket: str,
    source_sheet_order: int,
) -> SheetRecord:
    class_label = clean_text(ws["C7"].value)
    student_name = clean_text(ws["C8"].value)
    father_name = clean_text(ws["C9"].value)
    academic_year = clean_text(ws["B6"].value)

    if not student_name:
        student_name = f"Unnamed student {source_sheet_order}"

    return SheetRecord(
        workbook_file=workbook_file,
        workbook_bucket=workbook_bucket,
        class_label=class_label,
        sheet_name=ws.title,
        academic_year=academic_year,
        student_id=f"{slugify(workbook_bucket)}-{slugify(ws.title)}",
        student_name=student_name,
        father_name=father_name,
        source_sheet_order=source_sheet_order,
        subjects=parse_subject_rows(ws),
        total_max_marks=to_number(ws.cell(23, 3).value),
        total_min_marks=to_number(ws.cell(23, 4).value),
        class_total=to_number(ws.cell(23, 10).value),
        class_percentage=ratio_to_percentage(ws.cell(24, 10).value),
        attendance_count=to_number(ws.cell(32, 10).value),
        attendance_out_of=extract_first_number(ws.cell(32, 11).value),
        attendance_percentage=ratio_to_percentage(ws.cell(33, 11).value),
        result=clean_text(ws.cell(17, 11).value),
        grade=clean_text(ws.cell(20, 11).value),
        position=clean_text(ws.cell(22, 11).value),
        notes=clean_text(ws.cell(38, 10).value),
    )


def collect_subject_names(records: list[SheetRecord]) -> list[str]:
    names: list[str] = []
    for record in records:
        for subject in record.subjects:
            if subject.name not in names:
                names.append(subject.name)
    return names


def build_import_rows(records: list[SheetRecord], subject_names: list[str]) -> list[list[str]]:
    if not records:
        return []

    first_lookup = {subject.name: subject for subject in records[0].subjects}
    rows: list[list[str]] = []

    for subject_name in subject_names:
        subject = first_lookup.get(subject_name)
        rows.append(
            [
                "__EDUSTATS_THRESHOLD__",
                subject_name,
                format_csv_value(subject.max_mark if subject else 100),
                format_csv_value(subject.min_mark if subject and subject.min_mark is not None else 33),
            ]
        )

    rows.append(["Student ID", "Full Name", *subject_names, "Mean"])

    for record in records:
        subject_lookup = {subject.name: subject for subject in record.subjects}
        marks: list[float] = []
        row = [record.student_id, record.student_name]
        for subject_name in subject_names:
            subject = subject_lookup.get(subject_name)
            value = subject.final_term if subject else None
            row.append(format_csv_value(value))
            if isinstance(value, (int, float)):
                marks.append(float(value))
        row.append(format_csv_value(round(mean(marks), 2) if marks else None))
        rows.append(row)

    return rows


def flatten_record(record: SheetRecord, subject_names: list[str]) -> dict[str, Any]:
    subject_lookup = {subject.name: subject for subject in record.subjects}
    flat: dict[str, Any] = {
        "workbook_file": record.workbook_file,
        "workbook_bucket": record.workbook_bucket,
        "class_label": record.class_label,
        "sheet_name": record.sheet_name,
        "academic_year": record.academic_year,
        "student_id": record.student_id,
        "student_name": record.student_name,
        "father_name": record.father_name,
        "source_sheet_order": record.source_sheet_order,
        "total_max_marks": record.total_max_marks,
        "total_min_marks": record.total_min_marks,
        "class_total": record.class_total,
        "class_percentage": record.class_percentage,
        "attendance_count": record.attendance_count,
        "attendance_out_of": record.attendance_out_of,
        "attendance_percentage": record.attendance_percentage,
        "result": record.result,
        "grade": record.grade,
        "position": record.position,
        "notes": record.notes,
    }

    for subject_name in subject_names:
        subject = subject_lookup.get(subject_name)
        prefix = f"subject::{subject_name}"
        flat[f"{prefix}::max"] = subject.max_mark if subject else None
        flat[f"{prefix}::min"] = subject.min_mark if subject else None
        flat[f"{prefix}::first_term"] = subject.first_term if subject else None
        flat[f"{prefix}::mid_term"] = subject.mid_term if subject else None
        flat[f"{prefix}::final_term"] = subject.final_term if subject else None
        flat[f"{prefix}::weighted_score"] = subject.weighted_score if subject else None
        flat[f"{prefix}::overall_score"] = subject.overall_score if subject else None
        flat[f"{prefix}::total"] = subject.total if subject else None

    return flat


def write_csv(path: Path, rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)


def write_dict_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: format_csv_value(value) for key, value in row.items()})


def extract_folder(input_dir: Path, output_dir: Path) -> dict[str, Any]:
    workbook_paths = sorted(input_dir.glob("*.xlsx"))
    output_dir.mkdir(parents=True, exist_ok=True)

    summary: dict[str, Any] = {
        "input_dir": str(input_dir),
        "output_dir": str(output_dir),
        "workbooks": [],
        "total_students": 0,
    }

    all_records: list[SheetRecord] = []

    for workbook_path in workbook_paths:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        records: list[SheetRecord] = []

        for index, sheet_name in enumerate(wb.sheetnames, start=1):
            record = parse_sheet(wb[sheet_name], workbook_path.name, workbook_path.stem, index)
            records.append(record)
            all_records.append(record)

        wb.close()

        subject_names = collect_subject_names(records)
        class_label = records[0].class_label if records else workbook_path.stem
        class_slug = slugify(class_label)
        bucket_slug = slugify(workbook_path.stem)
        class_dir = output_dir / f"{bucket_slug}__{class_slug}"

        write_csv(class_dir / "edustats-import.csv", build_import_rows(records, subject_names))

        raw_rows = [flatten_record(record, subject_names) for record in records]
        if raw_rows:
            write_dict_csv(class_dir / "students-raw.csv", raw_rows, list(raw_rows[0].keys()))

        summary["workbooks"].append(
            {
                "workbook_file": workbook_path.name,
                "workbook_bucket": workbook_path.stem,
                "class_label": class_label,
                "class_slug": class_slug,
                "student_count": len(records),
                "subjects": subject_names,
                "output_folder": str(class_dir),
            }
        )
        summary["total_students"] += len(records)

    summary["all_subjects"] = collect_subject_names(all_records)
    (output_dir / "summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")

    manifest_rows = [
        {
            "workbook_file": item["workbook_file"],
            "class_label": item["class_label"],
            "student_count": item["student_count"],
            "subjects": "; ".join(item["subjects"]),
            "output_folder": item["output_folder"],
        }
        for item in summary["workbooks"]
    ]
    if manifest_rows:
        write_dict_csv(
            output_dir / "manifest.csv",
            manifest_rows,
            ["workbook_file", "class_label", "student_count", "subjects", "output_folder"],
        )

    return summary


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Extract class-grouped test data from Excel workbooks.")
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR, help="Folder containing .xlsx files")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Folder to write CSV outputs")
    return parser


def main() -> int:
    args = build_arg_parser().parse_args()
    summary = extract_folder(args.input_dir, args.output_dir)
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())